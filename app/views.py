from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from app.models import login , utype, department, course, assignment, submission,studymaterial, attstring, teacher, application, parent, record,attendence, attendancepercent, student, batch, subject, Room, time_slots, DAYS_OF_WEEK, MeetingTime
from django.contrib import messages
from django.core.mail import send_mail
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
import cms.settings
from django.contrib.auth.models import User
from django.views.decorators.cache import cache_control
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
import openpyxl
import json
from datetime import *
from matplotlib import pyplot as plt
import pytz
from django.utils.dateparse import parse_date

utc=pytz.UTC

#------------------------------
import random as rnd
POPULATION_SIZE = 6
NUMB_OF_ELITE_SCHEDULES = 1
TOURNAMENT_SELECTION_SIZE = 3
MUTATION_RATE = 0.05
D = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
HOURS=[1,2,3,4,5,6]
#------------------------------
class Data:
    def __init__(self):
        self._rooms = Room.objects.all()
        self._meetingTimes = MeetingTime.objects.all()
        self._teachers = teacher.objects.all()
        self._subjects = subject.objects.all()
        self._courses = course.objects.all()

    def get_rooms(self): return self._rooms

    def get_teachers(self): return self._teachers

    def get_subjects(self): return self._subjects

    def get_courses(self): return self._courses

    def get_meetingTimes(self): return self._meetingTimes


class Schedule:
    def __init__(self):
        self._data = data
        self._classes = []
        self._numberOfConflicts = 0
        self._fitness = -1
        self._classNumb = 0
        self._isFitnessChanged = True

    def get_classes(self):
        self._isFitnessChanged = True
        return self._classes

    def get_numbOfConflicts(self): return self._numberOfConflicts

    def get_fitness(self):
        if self._isFitnessChanged:
            self._fitness = self.calculate_fitness()
            self._isFitnessChanged = False
        return self._fitness

    def initialize(self):
        occupied_slots = {}
        batchs = batch.objects.all()
        for batche in batchs:
            course = batche.course
            n = batche.num_class_in_week         
            if n <= len(MeetingTime.objects.all()):
                subjects = course.subjects.all()
                for subject in subjects:
                    for i in range(n // len(subjects)):
                        crs_inst = subject.teachers.all()
                        newClass = Class(self._classNumb, course, batche.batch_id, subject)
                        self._classNumb += 1
                        a = data.get_meetingTimes()[rnd.randrange(0, len(MeetingTime.objects.all()))]
                        newClass.set_meetingTime(a)                         
                        newClass.set_room(data.get_rooms()[rnd.randrange(0, len(data.get_rooms()))])
                        newClass.set_teacher(crs_inst[rnd.randrange(0, len(crs_inst))])
                        self._classes.append(newClass)
            else:
                n = len(MeetingTime.objects.all())
                subjects = course.subjects.all()
                for subject in subjects:
                    for i in range(n // len(subjects)):
                        crs_inst = subject.teachers.all()
                        newClass = Class(self._classNumb, course, batche.batch_id, subject)
                        self._classNumb += 1
                        newClass.set_meetingTime(data.get_meetingTimes()[rnd.randrange(0, len(MeetingTime.objects.all()))])

                

                        newClass.set_room(data.get_rooms()[rnd.randrange(0, len(data.get_rooms()))])
                        newClass.set_teacher(crs_inst[rnd.randrange(0, len(crs_inst))])
                        self._classes.append(newClass)

              

        return self

    def calculate_fitness(self):
        self._numberOfConflicts = 0
        classes = self.get_classes()
        #OBTAIN teacher DICT FOR HOUR CALCULATION
        dat=Data()
        teachers = dat.get_teachers()
        x = (teachers[l].uid for l in range(len(teachers)))
        y = 0
        hour_count = dict.fromkeys(x, y)
       

        for i in range(len(classes)):
            if classes[i].room.seating_capacity < int(classes[i].subject.max_numb_students):
                self._numberOfConflicts += 1
            
            for x,y in hour_count.items():
                if classes[i].teacher.uid == x:
                    hour_count[x]+=1    

            for j in range(len(classes)):
                if j >= i:
                    if (classes[i].meeting_time == classes[j].meeting_time) and \
                            (classes[i].batch_id != classes[j].batch_id) and (classes[i].batch == classes[j].batch):
                        if classes[i].room == classes[j].room:
                            self._numberOfConflicts += 1
                        if classes[i].teacher == classes[j].teacher:
                            self._numberOfConflicts += 1

                            
        return 1 / (1.0 * self._numberOfConflicts + 1)


class Population:
    def __init__(self, size):
        self._size = size
        self._data = data
        self._schedules = [Schedule().initialize() for i in range(size)]

    def get_schedules(self):
        return self._schedules


class GeneticAlgorithm:
    def evolve(self, population):
        return self._mutate_population(self._crossover_population(population))

    def _crossover_population(self, pop):
        crossover_pop = Population(0)
        for i in range(NUMB_OF_ELITE_SCHEDULES):
            crossover_pop.get_schedules().append(pop.get_schedules()[i])
        i = NUMB_OF_ELITE_SCHEDULES
        while i < POPULATION_SIZE:
            schedule1 = self._select_tournament_population(pop).get_schedules()[0]
            schedule2 = self._select_tournament_population(pop).get_schedules()[0]
            crossover_pop.get_schedules().append(self._crossover_schedule(schedule1, schedule2))
            i += 1
        return crossover_pop

    def _mutate_population(self, population):
        for i in range(NUMB_OF_ELITE_SCHEDULES, POPULATION_SIZE):
            self._mutate_schedule(population.get_schedules()[i])
        return population

    def _crossover_schedule(self, schedule1, schedule2):
        crossoverSchedule = Schedule().initialize()
        for i in range(0, len(crossoverSchedule.get_classes())):
            if rnd.random() > 0.5:
                crossoverSchedule.get_classes()[i] = schedule1.get_classes()[i]
            else:
                crossoverSchedule.get_classes()[i] = schedule2.get_classes()[i]
        return crossoverSchedule

    def _mutate_schedule(self, mutateSchedule):
        schedule = Schedule().initialize()
        for i in range(len(mutateSchedule.get_classes())):
            if MUTATION_RATE > rnd.random():
                mutateSchedule.get_classes()[i] = schedule.get_classes()[i]
        return mutateSchedule

    def _select_tournament_population(self, pop):
        tournament_pop = Population(0)
        i = 0
        while i < TOURNAMENT_SELECTION_SIZE:
            tournament_pop.get_schedules().append(pop.get_schedules()[rnd.randrange(0, POPULATION_SIZE)])
            i += 1
        tournament_pop.get_schedules().sort(key=lambda x: x.get_fitness(), reverse=True)
        return tournament_pop


class Class:
    def __init__(self, id, course, batch, subject):
        self.batch_id = id
        self.course = course
        self.subject = subject
        self.teacher = None
        self.meeting_time = None
        self.room = None
        self.batch = batch

    def get_id(self): return self.batch_id

    def get_course(self): return self.course

    def get_subject(self): return self.subject

    def get_teacher(self): return self.teacher

    def get_meetingTime(self): return self.meeting_time

    def get_room(self): return self.room

    def set_teacher(self, teacher): self.teacher = teacher

    def set_meetingTime(self, meetingTime): self.meeting_time = meetingTime

    def set_room(self, room): self.room = room


data = Data()


def context_manager(schedule):
    classes = schedule.get_classes()
    context = []
    cls = {}
    for i in range(len(classes)):
        cls["batch"] = classes[i].batch_id
        cls['course'] = classes[i].course.course_name
        cls['subject'] = f'{classes[i].subject.subject_name} ({classes[i].subject.subject_number}, ' \
                        f'{classes[i].subject.max_numb_students}'
        cls['room'] = f'{classes[i].room.r_number} ({classes[i].room.seating_capacity})'
        cls['teacher'] = f'{classes[i].teacher.name} ({classes[i].teacher.uid})'
        cls['meeting_time'] = [classes[i].meeting_time.pid, classes[i].meeting_time.day, classes[i].meeting_time.time]
        context.append(cls)
    return context


def home(request):
    return render(request, 'index.html', {})


def timetable(request):
    schedule = []
    population = Population(POPULATION_SIZE)
    generation_num = 0
    population.get_schedules().sort(key=lambda x: x.get_fitness(), reverse=True)
    geneticAlgorithm = GeneticAlgorithm()
    while population.get_schedules()[0].get_fitness() != 1.0:
        generation_num += 1
        print('\n> Generation #' + str(generation_num))
        population = geneticAlgorithm.evolve(population)
        population.get_schedules().sort(key=lambda x: x.get_fitness(), reverse=True)
        schedule = population.get_schedules()[0].get_classes()

    wb = Workbook()
    batches=batch.objects.all()
    hoursss=MeetingTime.objects.all()

    # dataa={}

    for b in batches:
        ws = wb.create_sheet(b.batch_id)
        # dataa[b.batch_id]=[]
        ws['A1']="DAY"
        i=0
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=6):
            for cell in row:
                cell.value=str(D[i])
                i+=1
        i=0
        for col in ws.iter_cols(min_col=2, max_col=7, max_row=1):
            for cell in col:
                cell.value=HOURS[i]
                i+=1        
    for classs in schedule:
        mysheet=wb[classs.batch]
        if(classs.meeting_time.pid) in range(1,7):
            mysheet.cell(row=2,column=classs.meeting_time.pid+1).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)
            # dataa[b.batch_id].append(classs.subject.subject_number)
        elif(classs.meeting_time.pid) in range(7,13): 
            if(classs.meeting_time.pid%6==0):
                mysheet.cell(row=3,column=7).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)   
                # dataa[b.batch_id].append(classs.subject.subject_number)
            else:
                mysheet.cell(row=3,column=(classs.meeting_time.pid%6)+1).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)
                # dataa[b.batch_id].append(classs.subject.subject_number)
        elif(classs.meeting_time.pid) in range(13,19):    
             if(classs.meeting_time.pid%6==0):
                 mysheet.cell(row=4,column=7).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)
                #  dataa[b.batch_id].append(classs.subject.subject_number)
             else:
                 mysheet.cell(row=4,column=(classs.meeting_time.pid%6)+1).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)
                #  dataa[b.batch_id].append(classs.subject.subject_number)
        elif(classs.meeting_time.pid) in range(19,25):    
            if(classs.meeting_time.pid%6==0):
                 mysheet.cell(row=5,column=7).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)
                #  dataa[b.batch_id].append(classs.subject.subject_number)
            else:
                 mysheet.cell(row=5,column=(classs.meeting_time.pid%6)+1).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)
                #  dataa[b.batch_id].append(classs.subject.subject_number)
        elif(classs.meeting_time.pid) in range(25,31):    
            if(classs.meeting_time.pid%6==0):
                 mysheet.cell(row=6,column=7).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)
                #  dataa[b.batch_id].append(classs.subject.subject_number)
            else:
                 mysheet.cell(row=6,column=(classs.meeting_time.pid%6)+1).value=str(classs.subject.subject_number)+","+str(classs.teacher.uid)  
                #  dataa[b.batch_id].append(classs.subject.subject_number)
    f=wb["Sheet"]
    wb.remove(f)
    wb.save('timetable\\timetable.xlsx') 
    
    wb = openpyxl.load_workbook('timetable\\timetable.xlsx')     
    batches = batch.objects.all()
    for bat in batches:
        schedule = {}
        schedule.update({bat.batch_id : []})
        mysheet=wb[bat.batch_id]
        for i in range(2,7):
            for j in range(2,8):
                x = mysheet.cell(row=i, column=j)
                schedule[bat.batch_id].append((x.value.split(','))[0])
    # data = {
    #     "timetable":schedule,
    #     "batch":batches,
    # }

    
    batches=batch.objects.all()
    for b in batches:
        try:
            a=attstring.objects.filter(batch_id=b.batch_id)
            for i in D:
                    a=attstring.objects.get(batch_id=b.batch_id,day=i)
                    a.def_string=""
                    a.save()
                    a=attstring.objects.get(batch_id=b.batch_id,day=i)
                    temp=[]
                    if i == "Monday":
                        for k in range(0,6):
                            temp.append(schedule[b.batch_id][k])
                    elif i=="Tuesday":
                        for k in range(6,12):
                            temp.append(schedule[b.batch_id][k])
                    elif i=="Wednesday":
                        for k in range(12,18):
                            temp.append(schedule[b.batch_id][k])
                    elif i=="Thursday":
                        for k in  range(18,24):
                            temp.append(schedule[b.batch_id][k])
                    elif i=="Friday":
                        for k in range(24,30):
                            temp.append(schedule[b.batch_id][k])       
                    for smh in range(len(temp)):
                        if smh==0:
                            a.def_string+=str(temp[smh])
                        else:
                            a.def_string+=','+str(temp[smh])
                    a.save()
        except attstring.DoesNotExist:
            for b in batches:
                for i in D:
                        a=attstring.objects.create(batch_id=b.batch_id,day=i)
                        temp=[]
                    
                        if i == "Monday":
                            for k in range(0,6):
                                temp.append(schedule[b.batch_id][k])
                        elif i=="Tuesday":
                            for k in range(6,12):
                                temp.append(schedule[b.batch_id][k])
                        elif i=="Wednesday":
                            for k in range(12,18):
                                temp.append(schedule[b.batch_id][k])
                        elif i=="Thursday":
                            for k in range(18,24):
                                temp.append(schedule[b.batch_id][k])
                        elif i=="Friday":
                            for k in range(24,30):
                                temp.append(schedule[b.batch_id][k])
                        a.def_string=""        
                        for smh in range(len(temp)):
                            if smh==0:
                                a.def_string+=str(temp[smh])
                            else:
                                a.def_string+=','+str(temp[smh])
                        a.save()    
            

    return render(request, 'timetable.html', {'schedule': schedule, 'batchs': batch.objects.all(),
                                              'times': MeetingTime.objects.all()})


#----------------------------------------------------------------------------------------------------------------------------------







@cache_control(no_cache=True, must_revalidate=True)
def func():
  #some code
  return

# Create your views here.
def home(request):
    return render(request,'home.html')

def depts(request):
    data=department.objects.all()
    dep={
        "dept_no":data
    }
    return render(request,'departments.html',dep) 

def courses(request):
    data=department.objects.all()
    data2=course.objects.all()
    d={
        "dept_data":data,
        "course_data":data2,
    }
    return render(request,'courses.html',d) 

def log_in(request):
    return render(request,'login.html')    

def user_login(request):
    dat=login.objects.all()
    dat2=utype.objects.all()
    flag=0
    un=request.POST.get("username")
    pw=request.POST.get("password")
    for d in dat:
        if d.username==un and d.password==pw  and d.status=='1':
            request.session['id']=d.id
            flag =1
            id = request.session['id']
            data=login.objects.get(id=id)
            for e in dat2:
                if d.utype_id==e.id:
                    request.session['utype']=d.utype_id
                    if e.name=='admin':
                        return redirect('/dash/')
                    elif e.name=='teacher':
                        dat3=teacher.objects.get(login_id=d.id)
                        if dat3.name=='no data':
                            return render(request, 'teacherregister.html')
                        else:    
                            return redirect('/dash2/')  
                    elif e.name=='hod':
                        return redirect('/dash3/')     
                    elif e.name=='admission':
                       return redirect('/dash4/')
                    elif e.name=='student':
                        return redirect('/dash5/')   
                    else:
                        return render(request, 'login.html')         
    if flag==0:
        messages.error(request,'Username or Password is incorrect!')
        return redirect('/login/')         

def test_form(request):
    return render(request, 'form-samples-copy.html')
def dash(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    return render(request, 'dashboard.html')  

def dash2(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    return render(request, 'teacher.html')   

def dash3(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    return render(request, 'hod.html')   

def dash4(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    return render(request, 'incharge.html') 

def dash5(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    return render(request, 'student.html') 

def dashboardref(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    uid=request.session.get('utype') 
    print(uid)
    if (uid == 1):   
        data=application.objects.all()
        return render(request, 'dashboard.html') 
    elif (uid == 2):
        return render(request, 'teacher.html')    
    elif (uid == 3):
        return render(request, 'hod.html')   
    elif (uid == 4):
        return render(request, 'incharge.html')    
    elif (uid == 5):  
        return render(request, 'student.html')        

def deptadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    return render(request, 'deptadd.html')    

def deptaddval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    name = request.POST['name'] 
    descrip = request.POST['descrip']
    data=department.objects.all()
    for i in data:
        if i.name == name:
            messages.warning(request, 'Department already exists... Insertion failed!')
            return redirect('/deptadd/')
    dept=department.objects.create(name=name,descrip=descrip)
    dept.save()
    messages.success(request, 'Department added successfully...!')
    return redirect('/deptadd/')

def deptview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data=department.objects.all()
    dept={
        "dept_no":data
    }
    return render(request,"deptview.html",dept)

def depteditview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data=department.objects.all()
    dept={
        "dept_no":data
    }
    return render(request,"depteditview.html",dept)    

def deptedit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data=department.objects.all()
    id=request.POST.get("id")
    for d in data:
        if int(id)==d.id:
          dept={"d":d}
          return render(request,"deptedit.html",dept)

def deptupdate(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get("id")
    data=department.objects.get(pk=id)
    data2=course.objects.filter(dept_id=id)
    data.name=request.POST.get("name")
    data.status=request.POST.get("status")
    data.descrip=request.POST.get("descrip")
    if data.status=='0':    
        for c in data2:
            c.status=data.status
            c.save()
    data.save()
    
    messages.success(request, 'Department updated successfully...!')
    return redirect("/depteditview/")

def logout(request):
    if request.session.is_empty():
        return HttpResponseRedirect('/login')
    request.session.flush()
    return HttpResponseRedirect('/login')

def courseadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data=department.objects.all()
    dept={
        "dept_no":data
    }
    return render(request,"courseadd.html",dept)  

def courseaddval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    name = request.POST['name'] 
    duration = request.POST['duration']
    dept_id = request.POST['dept_id']  
    fee = request.POST['fee']
    data2 = course.objects.all()
    for i in data2:
        if i.course_name == name:
            messages.warning(request, 'Course already exists..! Insertion failed!')
            return redirect('/courseadd/')

    c=course.objects.create(course_name=name,duration=duration,dept_id=dept_id,fee=fee)
    c.save()
    messages.success(request, 'Course added successfully...!')
    return redirect('/courseadd/')    
    
def courseview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data1=course.objects.all()
    data2=department.objects.all()
    data={
        "course":data1,
        "dept":data2
    }
    

    return render(request,"courseview.html",data)    

def courseeditview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data1=course.objects.all()
    data2=department.objects.all()
    data={
        "course":data1,
        "dept":data2
    }
    

    return render(request,"courseeditview.html",data)       

def courseedit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data=course.objects.all()
    data2=department.objects.all()
    cid=request.POST.get("cid")
    for c in data:
        if int(cid)==c.id:
          dat={"c":c,"d":data2}
          return render(request,"courseedit.html",dat)

def courseupdate(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get("id")
    data=course.objects.get(pk=id)
    data.name=request.POST.get("name")
    data.duration=request.POST.get("duration")
    data.fee = request.POST['fee']
    data.dept_id=request.POST.get("dept")
    data.status=request.POST.get("status")
    data.save()
    messages.success(request, 'Course updated successfully...!')
    return redirect("/courseeditview/")

def useradd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    return render(request, 'useradd.html')    

def teacheradd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data=department.objects.all()
    password = User.objects.make_random_password(length=14, allowed_chars="abcdefghjkmnpqrstuvwxyz01234567889") 
    dept={
        "dept_no":data,
        "password":password
    }
    return render(request, 'teacheradd.html',dept)

def admissionadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    password = User.objects.make_random_password(length=14, allowed_chars="abcdefghjkmnpqrstuvwxyz01234567889") 
    dat={
        "password":password
    }
    return render(request, 'admissionadd.html',dat)

def admissiongen(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    email = request.POST['email'] 
    password = request.POST['password']
    mailto = request.POST['mailto']
    data=login.objects.all()
    for d in data:
        if d.username == email:
            messages.warning(request, 'User already exists...!')
            return redirect('/useradd/')
    l=login.objects.create(username=email,password=password,utype_id=4,status='1')
    subject = 'Login credentials'
    message = f'Welcome to Smart Academic Scheduler. Here are your login credentials to manage student admission.\nUsername: {email}\nPassword: {password}'
    email_from = cms.settings.EMAIL_HOST_USER
    recipient_list = [mailto]
    send_mail( subject, message, email_from, recipient_list )
    l.save()
    login_id=l.id
    messages.success(request, 'Admission in-charge added successfully...!')
    return redirect('/useradd/')    

def teachergen(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    email = request.POST['email'] 
    password = request.POST['password']
    dept_id = request.POST['dept_id']  
    uid = request.POST['uid']
    data=login.objects.all()
    for d in data:
        if d.username == email:
            messages.warning(request, 'User already exists...!')
            return redirect('/useradd/')
    l=login.objects.create(username=email,password=password,utype_id=2,status='1')
    l.save()
    login_id=l.id
    t=teacher.objects.create(name='no data',phone='no data',dob=None,gender='no data',address='no data',email=email,dept_id=dept_id,qualification='no data',login_id=login_id,uid=uid)
    t.save()
    subject = 'Login credentials'
    message = f'Welcome to Smart Academic Scheduler. Here are your login credentials.\nUsername: {email}\nPassword: {password}'
    email_from = cms.settings.EMAIL_HOST_USER
    recipient_list = [email]
    send_mail( subject, message, email_from, recipient_list )
    messages.success(request, 'Teacher added successfully...!')
    return redirect('/useradd/')    

def userview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data1=login.objects.filter(~Q(utype_id='1'))
    data2=utype.objects.all()
    data={
        "login":data1,
        "utype":data2
    }
    return render(request, 'userview.html',data)

def usereditview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data1=login.objects.filter(~Q(utype_id='1'))
    data2=utype.objects.all()
    data={
        "login":data1,
        "utype":data2
    }
    return render(request, 'usereditview.html',data)        

def useredit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get('id')
    data1=login.objects.get(pk=id)    
    data2=utype.objects.all()
    data={
          'user':data1,
          'utype':data2,  
            }
    return render(request, 'useredit.html',data)

def userupdate(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get('id')
    data1=login.objects.get(pk=id)    
    data1.username=request.POST.get('username')
    data1.password=request.POST.get('password')
    data1.utype_id=request.POST.get('utype')   
    data1.status=request.POST.get('status')
    data1.save()
    return redirect('/usereditview/')

def teacherregister(request):
    uid=request.session['id']
    data=teacher.objects.get(login_id=uid)
    data.name=request.POST.get("name")
    data.address=request.POST.get("address")
    data.gender=request.POST.get("gender")
    data.dob=request.POST.get("dob")
    data.qualification=request.POST.get("qualification")
    data.phone=request.POST.get("phone")

    Photo=request.FILES['photo']
    fs=FileSystemStorage()
    fn=fs.save(Photo.name, Photo)
    uploaded_file_url=fs.url(fn)
    data.photo=uploaded_file_url

    data.save()
    return redirect('/login/')

def sethod(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    data2=department.objects.all()
    dat={
        
        "department":data2,
    }
    return render(request, 'sethod.html', dat)    

def deptteacherview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get("id") 
    data=teacher.objects.filter(dept_id=id) 
    dat={
        "teacher":data,     
    } 
    return render(request,'deptteacherview.html',dat)

def hodupdate(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get("id")
    hod=teacher.objects.get(pk=id)
    teachers=teacher.objects.filter(dept_id=hod.dept_id)
    loginfo=login.objects.filter(utype_id='3')
    for t in teachers:
        for l in loginfo:
            if t.login_id == l.id:
                l.utype_id='2'
                l.save()
    hodlogin=login.objects.get(pk=hod.login_id) 
    hodlogin.utype_id='3'
    hodlogin.save()           
    return redirect('/sethod/')

def hodview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    l=login.objects.filter(utype_id='3')
    t=teacher.objects.all()
    d=department.objects.all()
    data={
        "loginfo":l,
        "teachers":t,
        "dept":d,
    }
    return render(request,'hodview.html',data)    

def appli(request):
    c=course.objects.filter(status='1')
    data={
        "course":c,
    }
    return render(request,'application.html',data)   

def apply(request):
    name=request.POST.get('name') 
    dob=request.POST.get('dob') 
    gender=request.POST.get('gender') 
    address=request.POST.get('address')
    phone=request.POST.get('phone')
    email=request.POST.get('email')
    course=request.POST.get('course')
    fname=request.POST.get('fname')
    mname=request.POST.get('mname')
    fmail=request.POST.get('fmail')
    mmail=request.POST.get('mmail')
    fjob=request.POST.get('fjob')
    mjob=request.POST.get('mjob')
    fphone=request.POST.get('fphone')
    mphone=request.POST.get('mphone')
    tenth=request.POST.get('tenth')
    twelfth=request.POST.get('twelfth')
    ug=request.POST.get('ug')
    
    data=application.objects.create(name=name,dob=dob,gender=gender,address=address,phone=phone,email=email,course_id=course)

    #Photo upload
    Photo=request.FILES['photo']
    fs=FileSystemStorage()
    fn=fs.save(Photo.name, Photo)
    uploaded_file_url=fs.url(fn)
    data.photo=uploaded_file_url
    data.save()


    #Certificate upload
    certificatetenth=request.FILES['certificatetenth']
    fs=FileSystemStorage()
    fn=fs.save(certificatetenth.name, certificatetenth)
    uploaded_file_url=fs.url(fn)
    certificatetenth=uploaded_file_url

    certificatetwelfth=request.FILES['certificatetwelfth']
    fs=FileSystemStorage()
    fn=fs.save(certificatetwelfth.name, certificatetwelfth)
    uploaded_file_url=fs.url(fn)
    certificatetwelfth=uploaded_file_url  

    certificateug=request.FILES['certificateug']
    fs=FileSystemStorage()
    fn=fs.save(certificateug.name, certificateug)
    uploaded_file_url=fs.url(fn)
    certificateug=uploaded_file_url  
    data3=record.objects.create(tenth=tenth,twelfth=twelfth,ug=ug,certificatetenth=certificatetenth,certificatetwelfth=certificatetwelfth,certificateug=certificateug,app_id=data.id)

    data3.save()

    data2=parent.objects.create(fname=fname,mname=mname,fmail=fmail,mmail=mmail,fjob=fjob,mjob=mjob,fphone=fphone,mphone=mphone,app_id=data.id)
    data2.save()
    
    return render(request,'home.html')

def courseapp(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    c=course.objects.all()
    d=department.objects.all()
    data={
        "course":c,
        "dept":d,
    }
    return render(request,'courseapp.html',data)

def appliview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    cid=request.POST.get("cid")
    cname=request.POST.get("cname")
    a=application.objects.filter(course_id=cid,stage__lt=2)
    r=record.objects.all()
    data={
        "app":a,
        "rec":r,
        "course":cname,
    }
    return render(request,'appliview.html',data)

def confirmapp(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get("aid")  
    a=application.objects.get(id=id)
    c=course.objects.get(id=a.course_id)
    subject = 'Application confirmation'
    message = f'Welcome to EduExpert.\nApplicant name: {a.name}\nPhone: {a.phone}\nCourse: {c.course_name}\n\n\n Please contact us to confirm your application. You can do so by replying to this email or contacting us at +919228833746.'
    email_from = cms.settings.EMAIL_HOST_USER
    recipient_list = [a.email]
    send_mail( subject, message, email_from, recipient_list )
    a.stage='1.25'
    a.save()
    snd1=application.objects.filter(course_id=c.id,stage__lt=2)
    snd2=record.objects.all()
    data={
        "app":snd1,
        "rec":snd2,
        "course":c.course_name,
    }
    return render(request,'appliview.html',data)

def confirmedapp(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.POST.get("aid")  
    a=application.objects.get(id=id)
    c=course.objects.get(id=a.course_id)
    a.stage='1.5'
    a.save()
    snd1=application.objects.filter(course_id=c.id,stage__lt=2)
    snd2=record.objects.all()
    data={
        "app":snd1,
        "rec":snd2,
        "course":c.course_name,
    }
    return render(request,'appliview.html',data)

def ranklistview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    c=course.objects.all()
    d=department.objects.all()
    data={
        "course":c,
        "dept":d,
    }
    return render(request,'ranklistview.html',data)    

def ranklist(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    cid=request.POST.get("cid")
    a=application.objects.filter(stage = '1.5',course_id=cid) 
    c=course.objects.get(id=cid) 
    r=record.objects.all()
    for i in a:
        for j in r:
            if i.id == j.app_id:
                if j.ug is not None:
                    i.score=(j.tenth + j.twelfth + j.ug)/3
                    i.save()
                else:    
                    i.score=((0.3*j.tenth+0.7*j.twelfth)/3)
                    i.save()
    b=application.objects.filter(stage = '1.5',course_id=cid).order_by('-score')     
    print(b)           
    data={
        "applicants":b,
        "course":c.course_name,
    }     
    return render(request,'ranklist.html',data)  

def sendinvite(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    applicants= request.POST.getlist('invite[]')
    if applicants==[]:
        return redirect('/ranklistview/') 
    cid=request.POST.get("cid")
    c=course.objects.get(id=cid)
    date=request.POST.get("date")  
    for i in applicants:
        a=application.objects.get(id=int(i))
        subject = 'Interview invitation'
        message = f'Welcome to EduExpert.\nApplicant name: {a.name}\nPhone: {a.phone}\nCourse: {c.course_name}\n\n\n The admission process has been scheduled on { date }. You are requested to reach the college by 9:00 am with the required documents.'
        email_from = cms.settings.EMAIL_HOST_USER
        recipient_list = [a.email]
        send_mail( subject, message, email_from, recipient_list )
        a.stage='2'
        a.save()
    return redirect('/ranklistview/')    

def courseapp2(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    c=course.objects.all()
    d=department.objects.all()
    data={
        "course":c,
        "dept":d,
    }
    return render(request,'courseapp2.html',data)

def appliview2(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    cid=request.POST.get("cid")
    cname=request.POST.get("cname")
    a=application.objects.filter(course_id=cid,stage='2')
    r=record.objects.all()
    data={
        "app":a,
        "rec":r,
        "course":cname,
    }
    return render(request,'appliview2.html',data)

def verify(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    cid=request.POST.get("cid")
    aid=request.POST.get("aid")
    app=application.objects.get(id=aid)
    #student admission
    s=student.objects.create(app_id=aid)
    s.save()
    app.stage="3"
    app.save()
    cname=request.POST.get("cname")
    a=application.objects.filter(course_id=cid,stage='2')
    r=record.objects.all()
    data={
        "app":a,
        "rec":r,
        "course":cname,
    }
    return render(request,'appliview2.html',data)

def newadmissions(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    h=teacher.objects.get(login_id=id)
    dept=h.dept_id
    c=course.objects.filter(dept_id=dept)        
    data={
        "course":c,
    }                
   
    return render(request,'newadmissions.html',data)

def newadmissionview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    cid=request.POST.get("cid")
    id=request.session.get("id")
    h=teacher.objects.get(login_id=id)
    dept=h.dept_id
    c=course.objects.filter(dept_id=dept)
    ss=student.objects.filter(batch_id=0)
    stud=[]
    for sss in ss:
        stud.append(sss.app_id)
    a=application.objects.filter(stage='3',course_id=cid,pk__in=stud)
    finalapps=[]
    cs=[]
    for app in a:
        for i in c:
            if app.course_id == i.id:
                finalapps.append(app.id)
               
    ap=application.objects.filter(pk__in=finalapps)                    
    data={
        "app":ap,
        "course":c,
        "cname":course.objects.get(id=cid),
        "batch":batch.objects.filter(course_id=cid)
    }                

    return render(request,'newadmissionview.html',data)    

def batchadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")    
    h=login.objects.get(id=id)
    hod=teacher.objects.get(login_id=id)
    dept_id=hod.dept_id
    c=course.objects.filter(dept_id=dept_id)
    l=login.objects.filter(utype_id__in=['2','3'],status='1')
    activeteachers=[]
    for ln in l:
        activeteachers.append(ln.id)       
    t=teacher.objects.filter(~Q(name='null'),dept_id=dept_id,login_id__in=activeteachers)
    data={
        "course":c,
        "teacher":t,
    }
    return render(request,'batchadd.html',data)   

def batchaddval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    batch_id=request.POST.get("batch_id")
    num_class_in_week=request.POST.get("num_class_in_week")
    class_teacher=request.POST.get("class_teacher")
    course_id=request.POST.get("course_id")
    s=batch.objects.create(course_id=course_id,class_teacher=class_teacher,num_class_in_week=num_class_in_week,batch_id=batch_id)
    s.save()
    return redirect('/batchadd/')

def batchview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    hod=teacher.objects.get(login_id=id)
    dept_id=hod.dept_id
    c=course.objects.filter(dept_id=dept_id)
    courses=[]
    for i in c:
        courses.append(i.id)
    batches=batch.objects.filter(course_id__in=courses)
    t=teacher.objects.filter(dept_id=dept_id)
    data={
        "batch":batches,
        "course":c,
        "teacher":t,
    }
    return render(request,'batchview.html',data)  

def addstudent(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    bid=request.POST.get("batch")
    sid= request.POST.getlist("students[]")
    if sid==[]:
        return redirect('/newadmissions/') 
    for s in sid:
        stud=student.objects.get(app_id=int(s))
        stud.batch_id=bid
        stud.save()
    return redirect('/newadmissions/')    

def subjectadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    hod=teacher.objects.get(login_id=id)
    dept=hod.dept_id
    data={
        "deptid":dept,
    }     
    return render(request,'subjectadd.html',data)      



def subjectaddval(request):
    subject_number=request.POST.get("subject_number")
    subject_name=request.POST.get("subject_name")
    max_numb_students=request.POST.get("max_numb_students")
    subject_type=request.POST.get("subject_type")
    dept_id = request.POST.get("dept_id")
    s=subject.objects.create(subject_number=subject_number,subject_name=subject_name,subject_type=subject_type,max_numb_students=max_numb_students,dept_id=dept_id)
    s.save()
    return redirect('/subjectadd')

def subjectview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    hod=teacher.objects.get(login_id=id)
    dept=hod.dept_id    
    sub=subject.objects.filter(dept_id=dept)
    c=course.objects.filter(dept_id=dept)
    data={
        "subjects":sub,
        "courses":c,
    }    
    return render(request,'subjectview.html',data)

def subjectedit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    s=request.POST["subject_number"]    
    sub=subject.objects.get(subject_number=s)
    data={
        "subject":sub,
    }    
    return render(request,'subjectedit.html',data)

def subjectupdateval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    subject_name=request.POST["subject_name"]
    subject_type=request.POST["subject_type"]
    max_numb_students=request.POST["max_numb_students"] 
    pk=request.POST["pk"]
    data=subject.objects.get(subject_number=pk)
    data.subject_name=subject_name
    data.subject_type=subject_type
    data.max_numb_students=max_numb_students
    data.save()
    return redirect('/subjectview')

def subjectaddcourse(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    hod=teacher.objects.get(login_id=id)
    dept=hod.dept_id
    c=course.objects.filter(dept_id=dept)
    s=subject.objects.filter(dept_id=dept)
    data={
        "courses":c,
        "subjects":s,
    }
    return render(request,'subjectaddcourse.html',data) 

def subjectaddcourseval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    course_id=request.POST.get("course_id")
    c=course.objects.get(id=course_id)
    subjects=request.POST.getlist("subject_id[]")
    
    for i in subjects:
        s=c.subjects.add(i)
    return redirect('/subjectaddcourse') 

def subjectcourseview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    c=course.objects.all()
    data={
        "courses":c,
    }    
    return render(request,'subjectcourseview.html',data)       

def subjectcourseedit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    cid=request.POST["cid"]    
    c=course.objects.get(id=cid)
    s=subject.objects.all()
    data={
        "courses":c,
        "subjects":s,
    }
    return render(request,'subjectcourseedit.html',data)

def subjectcourseupdateval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    cid=request.POST.get("cid")
    c=course.objects.get(id=cid)
    subj=request.POST.getlist("subject_id[]")
    sub=c.subjects.all()
    print(sub)
    for s in sub:
        if s.subject_number not in subj:                                                                                    
            s=c.subjects.remove(s)
    return redirect('/subjectcourseview')

def roomadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    return render(request,'roomadd.html') 

def roomview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    r = Room.objects.all() 
    data = {
        "rooms" : r,
    }   
    return render(request,'roomview.html',data)

def roomedit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id = request.POST["id"]    
    r = Room.objects.get(id=id)
    data = {
        "room" : r,
    }
    return render(request,'roomeditview.html',data)

def roomupdateval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id = request.POST["id"]
    room = Room.objects.get(id=id)
    room.r_number = request.POST["r_number"]
    room.seating_capacity = request.POST["seating_capacity"]
    room.save()
    return redirect('/roomview')

def roomaddval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    r_number=request.POST.get("r_number")
    seating_capacity=request.POST.get("seating_capacity")
    s=Room.objects.create(r_number=r_number,seating_capacity=seating_capacity)
    s.save()
    return redirect('/roomadd')    

def meetingadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    hr=[]
    day=[]
    for i in time_slots:
        hr.append(i[0])  
    for i in DAYS_OF_WEEK:
        day.append(i[0])     
    data={
        "hours":hr,
        "da":day,
    }
    return render(request, 'meetingadd.html', data)

def meetingview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    meetings = MeetingTime.objects.all()        
    data={
        "hours":meetings,
     
    }
    return render(request, 'meetingview.html', data)

def meetingedit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    pid=request.POST["pid"]    
    meeting = MeetingTime.objects.get(pid=pid) 
    hr=[]
    day=[]
    for i in time_slots:
        hr.append(i[0])  
    for i in DAYS_OF_WEEK:
        day.append(i[0])   
    data = {
        "hour":meeting,
        "days":day,
        "hours":hr,
        
    }
    return render(request, 'meetingeditview.html', data)

def meetingupdateval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    pid=request.POST["pid"] 
    meeting = MeetingTime.objects.get(pid=pid) 
    meeting.time = request.POST["time"]  
    meeting.day = request.POST["day"]
    meeting.save()
    return redirect('/meetingview')  

def meetingaddval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    pid=request.POST.get("pid")
    time=request.POST.get("time")
    day=request.POST.get("day")
    s=MeetingTime.objects.create(pid=pid, time=time, day=day)
    s.save()
    return redirect('/meetingadd')

def subjectaddteacher(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    hod=teacher.objects.get(login_id=id)
    dept_id=hod.dept_id
    t=teacher.objects.filter(dept_id=dept_id)
    cid=request.POST.get("cid")    
    c=course.objects.get(id=cid)
    sub=c.subjects.all()
      
    data={
        "subjects":sub,
        "teachers":t,
    }
    print(data)
    return render(request,"subjectaddteacher.html",data)           

def subjectaddteacherval(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    subject_number=request.POST.get("subject_number")
    teacher_id=request.POST.getlist("teacher_id[]")
    s=subject.objects.get(subject_number=subject_number)
    for i in teacher_id:
        c=s.teachers.add(i)
    return redirect('/batchview')

def subjectteacherview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')    
    id=request.session.get("id")
    t = teacher.objects.get(login_id=id)
    sub = subject.objects.filter(dept_id=t.dept_id)
    tea = teacher.objects.filter(dept_id=t.dept_id)
    data = {
        "subjects":sub,
        "teachers":tea,
    }
    return render(request,"subjectteacherview.html",data)

def subjectteacheredit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')    
    tid = request.POST["tid"]
    t = teacher.objects.get(id=tid)
    sub = subject.objects.filter(dept_id=t.dept_id)
    data = {
        "subjects" : sub,
        "teacher" : t,
    }
    return render(request,"subjectteacheredit.html",data)

def timetablegen(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    return render(request,"timetablegen.html")

def profile(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t = teacher.objects.get(login_id=id)
    d = department.objects.get(id=t.dept_id)
    data = {
        "hod" : t,
        "dept" : d,
    }
    return render(request,"profile.html",data)

def profileupdate(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")    
    t = teacher.objects.get(login_id=id)
    t.name = request.POST["name"]
    t.dob = request.POST["dob"]
    t.gender = request.POST["gender"]
    t.address = request.POST["address"]
    t.email = request.POST["email"]
    t.phone = request.POST["phone"]
    t.qualification = request.POST["qualification"]
    if 'photo' not in request.FILES:
        t.save()
    else:
        Photo=request.FILES['photo']
        fs=FileSystemStorage()
        fn=fs.save(Photo.name, Photo)
        uploaded_file_url=fs.url(fn)
        t.photo=uploaded_file_url   
        t.save() 
    return redirect('/profile')

def teacherview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session['id']
    d=teacher.objects.get(id=id)
    dep=teacher.objects.filter(dept_id=d.dept_id)
    data={'teacher':dep,}
    return render(request, 'teacherview.html',data)

def teacherprofile(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session['id']
    t=teacher.objects.get(login_id=id)
    data={"t":t}
    return render(request, 'teacherprofile.html',data)

def teacherupdate(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")    
    t = teacher.objects.get(login_id=id)
    t.name = request.POST["name"]
    t.dob = request.POST["dob"]
    t.gender = request.POST["gender"]
    t.address = request.POST["address"]
    t.email = request.POST["email"]
    t.phone = request.POST["phone"]
    t.qualification = request.POST["qualification"]
    if 'photo' not in request.FILES:
        t.save()
    else:
        Photo=request.FILES['photo']
        fs=FileSystemStorage()
        fn=fs.save(Photo.name, Photo)
        uploaded_file_url=fs.url(fn)
        t.photo=uploaded_file_url   
        t.save() 
    return redirect('/teacherprofile')


def timetableview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    wb = openpyxl.load_workbook('timetable\\timetable.xlsx')     
    batches = batch.objects.all()
    bat = request.POST.get("bid", batches[0].batch_id)
    schedule = {}
    schedule.update({bat : []})
    mysheet=wb[bat]
    for i in range(2,7):
        for j in range(1,8):
            x = mysheet.cell(row=i, column=j)
            schedule[bat].append(x.value)
    data = {
        "timetable":schedule,
        "batch":batches,
    }            
    print(data)
    return render(request,'timetableview.html',data)

def studentview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    try:
        b=batch.objects.get(class_teacher=t.id)
        s=student.objects.filter(batch_id=b.batch_id)
        list=[]
        for i in s:
            a=application.objects.get(id=i.app_id)
            list.append(a)
        print(list)
    except batch.DoesNotExist:
        return redirect('/teacher404')    
    return render(request,'studentview.html',{'l':list})

def rejectapp(request,id):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    a=application.objects.get(id=id)
    a.stage='rejected'
    a.save()
    return redirect('/appliview')

def rejectapp1(request,id):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    a=application.objects.get(id=id)
    a.stage='rejected'
    a.save()
    return redirect('/appliview2')

def teacher_timeview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    id=request.session['id']
    wb = openpyxl.load_workbook('timetable\\timetable.xlsx') 
    # izumi path = D:\Main Project\cms\\timetable\\timetable.xlsx
    # Adarsh path=D:\python\Smart Academic Scheduler\SmartAcademicScheduler\timetable    
    batches = batch.objects.all()
    bat = request.POST.get("bid", batches[0].batch_id)
    schedule = {}
    schedule.update({bat : []})
    mysheet=wb[bat]
    for i in range(2,7):
        for j in range(1,8):
            x = mysheet.cell(row=i, column=j)
            schedule[bat].append(x.value)
    t=teacher.objects.get(login_id=id)
    uid=t.uid 
    data = {
        "timetable":schedule,
        "batch":batches,
        "uid":uid,
        "d":D,
    }           
    print(data)
    return render(request, 'teacher_timeview.html',data)
def studentedit(request,id):
    a=application.objects.get(id=id)
    r=record.objects.get(app_id=a.id)
    data={"a":a,"r":r}
    return render(request, 'studentedit.html',data)

def studentupdate(request):
    id=request.POST.get('sid')
    s=application.objects.get(id=id)
    r=record.objects.get(app_id=id)
    s.name=request.POST.get('name')
    s.dob=request.POST.get('dob')
    s.gender=request.POST.get('gender')
    s.address=request.POST.get('address')
    s.phone=request.POST.get('phone')
    s.email=request.POST.get('email')
    if 'photo' in request.FILES:
        Photo=request.FILES['photo']
        fs=FileSystemStorage()
        fn=fs.save(Photo.name, Photo)
        uploaded_file_url=fs.url(fn)
        s.photo=uploaded_file_url
    if 'sslc' in request.FILES:
        sslc=request.FILES['sslc']
        fs=FileSystemStorage()
        fn=fs.save(sslc.name, sslc)
        uploaded_file_url=fs.url(fn)
        r.certificatetenth=uploaded_file_url
        msslc = request.POST.get('msslc')
        mplustwo = request.POST.get('mplustwo')
        mug = request.POST.get('mug')
        r.tenth=msslc
        r.twelfth=mplustwo
        r.ug=mug
        avg=(float(msslc)+float(mplustwo)+float(mug))/3
        s.score=avg
    if 'plustwo' in request.FILES:
        plustwo=request.FILES['plustwo']
        fs=FileSystemStorage()
        fn=fs.save(plustwo.name, plustwo)
        uploaded_file_url=fs.url(fn)
        r.certificatetwelfth=uploaded_file_url
        msslc = request.POST.get('msslc')
        mplustwo = request.POST.get('mplustwo')
        mug = request.POST.get('mug')
        r.tenth=msslc
        r.twelfth=mplustwo
        r.ug=mug
        avg=(float(msslc)+float(mplustwo)+float(mug))/3
        s.score=avg
    if 'ug' in request.FILES:
        mug=request.FILES['ug']
        fs=FileSystemStorage()
        fn=fs.save(mug.name, mug)
        uploaded_file_url=fs.url(fn)
        r.certificateug=uploaded_file_url
        msslc = request.POST.get('msslc')
        mplustwo = request.POST.get('mplustwo')
        mug = request.POST.get('mug')
        r.tenth=msslc
        r.twelfth=mplustwo
        r.ug=mug
        avg=(float(msslc)+float(mplustwo)+float(mug))/3
        s.score=avg
    s.save()
    r.save()
    return redirect('/studentview')

def attendenceview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    try:
        id=request.session.get("id")
        t=teacher.objects.get(login_id=id)
        b=batch.objects.get(class_teacher=t.id)
        s=student.objects.filter(batch_id=b.batch_id)
        list=[]
        for i in s:
            a=application.objects.get(id=i.app_id)
            list.append(a)
    except batch.DoesNotExist:  
        return HttpResponseRedirect('/teacher404')     
    return render(request,'attendenceview.html',{"l":list,"s":s})

def attendancemark(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session['id']
    T=teacher.objects.get(login_id=id)
    b=batch.objects.get(class_teacher=T.id)
    batchid=b.batch_id
    print(batchid)
    day=request.POST.get('cday')
    defstring=attstring.objects.get(batch_id=batchid,day=day)
    temp=defstring.def_string
    att_string=temp.split(",")
    print(att_string)
    students=request.POST.getlist('studentid[]')
    hours=request.POST.getlist('hour[]')
    print(students,hours)
    date=request.POST.get('date')
    for s in students:
        temp=str(s)+"-"
        temp2=""
        for h in hours:
            if temp in h:
                if '1hr' in h:
                    temp2+='1-'+str(att_string[0])+','
                elif '2hr' in h:
                    temp2+='2-'+str(att_string[1])+','
                elif '3hr' in h:
                    temp2+='3-'+str(att_string[2])+','
                elif '4hr' in h:
                    temp2+='4-'+str(att_string[3])+','
                elif '5hr' in h:
                    temp2+='5-'+str(att_string[4])+','
                elif '6hr' in h:
                    temp2+='6-'+str(att_string[5])+','
        try:
            a=attendence.objects.get(date=date,student_id=s)
            a.att_str=temp2
            a.save()
        except attendence.DoesNotExist:
            a=attendence()
            a.att_str=temp2
            a.student_id=s
            a.day=day
            a.date=date
            a.save()
        print(temp2)
    
    return redirect('/attendenceview')

def view_att(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    if request.method == 'POST':
        date=json.loads(request.body).get('date')
        id=request.session['id']
        T=teacher.objects.get(login_id=id)
        b=batch.objects.get(class_teacher=T.id)
        batchid=b.batch_id
        students=student.objects.filter(batch_id=batchid)
        a=attendence.objects.filter(date=date)
        att={}
        # for s in students:
        #     att[s.id] = []
        #     attobj = attendence.objects.get(student_id=s.id,date=date)
        #     temp=attobj.att_str
        #     if '1-' in str(temp):
        #         att[s.id].append('1')
        #     if '2-' in str(temp):
        #         att[s.id].append('2')
        #     if '3-' in str(temp):
        #         att[s.id].append('3')
        #     if '4-' in str(temp):
        #         att[s.id].append('4')
        #     if '5-' in str(temp):
        #         att[s.id].append('5')
        #     if '6-' in str(temp):
        #         att[s.id].append('6') 
        # print(att)                           
        data=a.values()
        return JsonResponse(list(data), safe=False)

def testmethod(request):
    
    
    working_days = data.objects.order_by().values('date').distinct()

    print(users)
    return render(request,'teacher404.html')

def teacher404(request):
    
    return render(request,'teacher404.html')

def timetableviewteacher(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)    
    wb = openpyxl.load_workbook('timetable\\timetable.xlsx')     
    try:
        batches = batch.objects.get(class_teacher=t.id)
        schedule = {}
        schedule.update({bat : []})
        mysheet=wb[bat]
        for i in range(2,7):
            for j in range(1,8):
                x = mysheet.cell(row=i, column=j)
                schedule[bat].append(x.value)
        data = {
            "timetable":schedule,
            "batch":batches,
        }
    except batch.DoesNotExist:  
        return HttpResponseRedirect('/teacher404')             
    return render(request,'timetableviewteacher.html',data)            

def studymaterialupload(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id) 
    sub=subject.objects.filter(teachers=t.id)
    data={
        "subjects":sub,
        "teacher":t,
    }

    return render(request,"studymaterial.html",data)

def studymaterialadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    teacher_id=t.id
    title=request.POST.get('title')
    subject_number=request.POST.get('subject_number')
    
    m=studymaterial()
    m.teacher_id=teacher_id
    m.title=title
    m.subject_number=subject_number

    mat=request.FILES['mate']
    fs=FileSystemStorage()
    fn=fs.save(mat.name, mat)
    uploaded_file_url=fs.url(fn)
    m.material=uploaded_file_url   
    m.save() 

    return redirect('/studymaterialupload')

def studymaterialview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    teacher_id=t.id
    m=studymaterial.objects.filter(teacher_id=teacher_id)
    s=subject.objects.filter(teachers=teacher_id)
    data={
        "materials" : m,
        "subjects" : s,
    }
    return render(request,'studymaterialview.html',data)

def studymaterialdelete(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    mid=request.POST.get('mid')
    m = studymaterial.objects.get(id = mid)
    m.delete()
    return redirect('/studymaterialview') 

def assignmentadd(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    teacher_id=t.id       
    s=subject.objects.filter(teachers=teacher_id)
    data={
        "subjects" : s,
        "teacher" : t,
    }
    return render(request,'assignmentadd.html',data)

def assignmentgen(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    teacher_id=t.id   
    subject_number = request.POST.get('subject_number')
    title = request.POST.get('title')
    problem = request.POST.get('problem')
    fromtime = request.POST.get('fromtime')
    totime = request.POST.get('totime')
    a = assignment.objects.create(teacher_id=teacher_id,fromtime=fromtime,totime=totime)
    a.subject_number = subject_number
    a.title = title
    a.problem = problem
    a.save()
    return redirect('/dashboardref')

def assignmentsubjects(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')        
    id = request.session.get("id")
    s = student.objects.get(login_id=id)    
    batche = batch.objects.get(batch_id=s.batch_id)
    coursee = course.objects.get(id=batche.course_id)
    subjects = coursee.subjects.all()
    data={
        "subjects" : subjects,
    }
    return render(request,'assignmentsubjects.html',data)

def assignments(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login') 
    sub_num = request.POST.get('subject_number')
    a = assignment.objects.filter(subject_number=sub_num,status=1)
    sub = subject.objects.get(subject_number=sub_num)
    t = teacher.objects.all()
    id = request.session.get("id")
    stud = student.objects.get(login_id=id)
    data={}
    data['submissions']=[]
    for assgnmt in a:       
        try:
            s = submission.objects.get(assignment_id=assgnmt.id,student_id=stud.id)
            data['submissions'].append(s)
        except submission.DoesNotExist:
            pass
    currenttime = utc.localize(datetime.now())
    
    data['assignments']=[]
    data['subject']=sub
    data['teachers']=t
    
    for assign in a:
        if assign.fromtime < currenttime and assign.totime > currenttime:
            data['assignments'].append(assign)        
        else:
            assign.status = 0   
            assign.save()

    return render(request,'assignments.html',data)

def assignmentsubmit(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    aid = request.POST.get('aid')    
    id = request.session.get("id")
    s = student.objects.get(login_id=id)    
    try:         
        sub = submission.objects.get(assignment_id=aid,student_id=s.id)
        ans=request.FILES['answer']
        fs=FileSystemStorage()
        fn=fs.save(ans.name, ans)
        uploaded_file_url=fs.url(fn)
        sub.answer=uploaded_file_url   
        sub.save()

    except submission.DoesNotExist:   
        sub = submission()
        sub.student_id=s.id
        sub.assignment_id=aid
        sub.marks=0
        ans=request.FILES['answer']
        fs=FileSystemStorage()
        fn=fs.save(ans.name, ans)
        uploaded_file_url=fs.url(fn)
        sub.answer=uploaded_file_url   
        sub.save()

    return redirect('/assignmentsubjects')

def assignmentview(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    tid=t.id 
    a = assignment.objects.filter(teacher_id=tid)
    print(a)
    s = subject.objects.all()
    data = {
        "assignments":a,
        "subjects":s,
    }

    return render(request,'assignmentview.html',data)

def assignmentsubmissions(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    aid = request.POST.get('aid') 
    a = assignment.objects.get(id=aid)
    sub = subject.objects.get(subject_number=a.subject_number)
    subm = submission.objects.filter(assignment_id=aid)
    stud = student.objects.all()
    app = application.objects.filter(stage=3)
    data = {
        "subject" : sub,
        "submissions" : subm,
        "assignment" : a,
        "students" : stud,
        "applications" : app,

    }    
    return render(request,'assignmentsubmissions.html',data)

def assignmentevaluate(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    sid = request.POST.get('sid') 
    marks = request.POST.get('marks') 
    s = submission.objects.get(id = sid)
    s.marks = marks
    s.save()
    return redirect('/assignmentview')

def assignmentresults(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id = request.session.get("id")
    s = student.objects.get(login_id=id)    
    batche = batch.objects.get(batch_id=s.batch_id)
    coursee = course.objects.get(id=batche.course_id)
    subjects = coursee.subjects.all()
    teachers = teacher.objects.all()
    data={}
    data["subjects"]=[]
    data["teachers"]=teachers
    data["subjects"].append(subjects)
    data["assignments"]=[]
    assign = assignment.objects.all()
    subm = submission.objects.filter(student_id=s.id)
    data["submissions"]=subm
    app = application.objects.get(id=s.app_id)
    data["student"]=app
    for assgn in assign:
        for sub in subjects:
            if assgn.subject_number == sub.subject_number:
                data["assignments"].append(assgn)

    return render(request,'assignmentresults.html',data)
def calattendence(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    b=batch.objects.get(class_teacher=t.id)    
    data={
        "batch":b.batch_id,
    }
    return render(request,'calattendence.html',data)   

def calatt(request):
    fdate=parse_date(request.POST.get('fdate'))
    tdate=parse_date(request.POST.get('tdate'))
    try:
        id=request.session.get("id")
        t=teacher.objects.get(login_id=id)
        b=batch.objects.get(class_teacher=t.id)
        s=student.objects.filter(batch_id=b.batch_id)
        data=[]
        dates=[]
        at=attendence.objects.all()
        for i in at:
            for j in s:
                if i.student_id == j.id and (i.date >= fdate and i.date <= tdate):
                    data.append(i)
        for k in data:
            dates.append(k.date)
        dates=set(dates)    
        dates=list(dates)
        c = course.objects.get(id=b.course_id)
        sub = c.subjects.all()
        hours={}
        attended={}
        for sb in sub:
            hours[sb.subject_number]=0
            attended[sb.subject_number]=0
        attstr = attstring.objects.filter(batch_id=b.batch_id)
        for a in dates:
            temp=[]
            for b in attstr:
                if(a.weekday()==0):
                    if b.day == "Monday":
                        temp = b.def_string.split(',')
                        for tmp in temp:
                            hours[tmp]+=1 
                elif(a.weekday()==1): 
                    if b.day == "Tuesday":
                        temp = b.def_string.split(',')  
                        for tmp in temp:
                            hours[tmp]+=1  
                elif(a.weekday()==2):
                    if b.day == "Wednesday":
                        temp = b.def_string.split(',') 
                        for tmp in temp:
                            hours[tmp]+=1   
                elif(a.weekday()==3):  
                    if b.day == "Thursday":
                        temp = b.def_string.split(',')
                        for tmp in temp:
                            hours[tmp]+=1 
                elif(a.weekday()==4):        
                    if b.day == "Friday":
                        temp = b.def_string.split(',')
                        for tmp in temp:
                            hours[tmp]+=1 
           
        for students in s:
            stud_attendance={}
            temp2=""
            for x in sub:
                stud_attendance[x.subject_number]=0 
            for attendancevalues in data:
                if students.id == attendancevalues.student_id:
                    temp = attendancevalues.att_str.split(',')
                    for i in temp:
                        for sbj in sub:
                            if sbj.subject_number in i:
                                stud_attendance[sbj.subject_number]+=1
                               
            print("student",stud_attendance)
            print("total",hours)
            for x in sub:
                temp2=temp2+str(x.subject_number)+"-"+str(round(((stud_attendance[x.subject_number]/hours[x.subject_number])*100),2))+","
            print(temp2)
            try:
                attper = attendancepercent.objects.get(student_id=students.id,fromdate=fdate,todate=tdate) 
                attper.attendancevalue = temp2
                attper.save()
            except attendancepercent.DoesNotExist:    
                attper = attendancepercent.objects.create(student_id=students.id,fromdate=fdate,todate=tdate,attendancevalue=temp2)      
                attper.save()
        return redirect('/calattendence')
    except batch.DoesNotExist:  
        return HttpResponseRedirect('/teacher404')             

def publishinternals(request):
    if request.session.is_empty():
        messages.error(request,'Session has expired, please login to continue!')
        return HttpResponseRedirect('/login')
    id=request.session.get("id")
    t=teacher.objects.get(login_id=id)
    try:
        b=batch.objects.get(class_teacher=t.id)
        data = {
            "batch":b.batch_id,
        }
        return render(request,'publishinternals.html',data)            
    except batch.DoesNotExist:
        return redirect('/teacher404')        